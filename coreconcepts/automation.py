#how to process spreadsheets through python
import openpyxl as xl
#imports 2 classes from the package
from openpyxl.chart import BarChart, Reference

#function to modify thousands of spreadsheets
def process_workbook(filename):
    #imports excel files with alias xl
    #loads the file name into a wb variable
    wb = xl.load_workbook(filename) #'transactions.xlsx'
    #case sensitive : returns a sheet from the xlsx by its name
    sheet = wb['Sheet1']

    #access a particular cell in the sheet
    #first method returns my coordinates. the second by row,column
    # cell = sheet['a1']
    # cell = sheet.cell(1,1)
    # print(cell.value)

    #iterate to get the value in 3rd column and multiply by 90% to complete the 5th column
    #how many rows does it has:
    #print(sheet.max_row)
    #range includes the numbers from the first to the last, but not the last, which is why we must include +1
    #first number is row start
    #for row in range(1, sheet.max_row + 1):
        #print(row)

    #starting in row 2, the values in column 3
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        #multiplies by 0.9
        corrected_price = cell.value * 0.9
        #includes the value into a new cell : cell object in the spreadsheet
        corrected_price_cell = sheet.cell(row,4)
        #add the value to the cell
        corrected_price_cell.value = corrected_price
        #print(cell.value)

    #selects from the second to the max row, values in the 4th col only. stored in values
    values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
    #creates an instance to the BarChart and fills it with the values
    chart = BarChart()
    chart.add_data(values)
    #adds the chart to the e2 coordinate top left
    sheet.add_chart(chart,'e2')
    #save into a new files(creates the file if it doesnt exist)
    wb.save(filename) #'transactions2.xlsx'


process_workbook('transactions.xlsx')


def test_workbook(file):
    wb = xl.load_workbook(file) #'transactions.xlsx'
    #case sensitive : returns a sheet from the xlsx by its name
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        #instancia dos celdas, donde sale el valor y donde lo guarda.
        cell = sheet.cell(row,3)
        corrected_price_cell = sheet.cell(row,15)
        #añade el valor a la celda
        corrected_price_cell.value = cell.value
    #elige el archivo donde hacer el cambio
    wb.save('transactions2.xlsx')


test_workbook('transactions.xlsx')